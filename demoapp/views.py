#Import Django modules
from django.shortcuts import render
from django.http import HttpResponse
from django.conf import settings

# Import Automation function modules
import os
from openpyxl import load_workbook
import win32com.client

#Import pythoncom
import pythoncom

# create a new instance of Word
word = win32com.client.Dispatch("Word.Application")

# Create your views here.
def index(request):
    return render(request, 'index.html')

def handle_uploaded_file(f, filename):
    file_path = os.path.join("demoapp", "SavedFiles", f.name)
    with open(file_path, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

def upload_files(request):
    if request.method == 'POST':
        # Get the uploaded files
        excel_file = request.FILES['excel_file']
        contract_file = request.FILES['contract_file']

        # Save the uploaded files to the desired directory
        handle_uploaded_file(excel_file, excel_file.name)
        handle_uploaded_file(contract_file, contract_file.name)

        # Call the Automation() function with the uploaded files
        Automation(excel_file.name,contract_file.name)

        # Get the file path of the generated contract
        contract_path = os.path.join("demoapp", "SavedFiles", 'New.docx')

        # Return a response that prompts the user to download the generated file
        with open(contract_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            response['Content-Disposition'] = 'attachment; filename="New.docx"'
            return response

    return render(request, 'index.html')

def Automation(excel_file,contract_file):
	# EDIT THE PARAMETERS BELOW!
    path = os.path.join("demoapp", "SavedFiles", contract_file)
    Workbook_name = os.path.join("demoapp", "SavedFiles", excel_file)
    Work_sheetname = "Input Sheet"
    Text_Replacement_Column = "C"
    Text_Target_Column = "D"

    # Call CoInitialize
    pythoncom.CoInitialize()

    #New_WordName

    wb = load_workbook(os.getcwd()+f"\\{Workbook_name}")  # Work Book
    ws = wb[Work_sheetname]  # Work Sheet Name
    column = ws[Text_Replacement_Column]  # Column to Extract Data From
    a = [column[x].value for x in range(len(column))]
    re_column=ws[Text_Target_Column]
    r=[re_column[x].value for x in range(len(re_column))]

    # Get the absolute path of the input file
    path = os.path.abspath(path)

    # Check if the input file exists
    if not os.path.exists(path):
        raise ValueError("File not found: " + path)

    # Open the Word document
    word = win32com.client.DispatchEx("Word.Application")
    document = word.Documents.Open(path)

    # Replace the old string with the new string
    for i in range(len(r)):
        word.Selection.Find.Execute(r[i], False, False, False, False, False, True, 1, True, a[i], 2)

    # Save and close the Word document
    try:
        document.SaveAs(os.path.join(settings.BASE_DIR, 'demoapp', 'SavedFiles', 'New.docx'))
        document.Close()
    except:
        print("Document Error Occured while saving")
        document.Close()

Automation("Book.xlsx","ITT Test.docx")